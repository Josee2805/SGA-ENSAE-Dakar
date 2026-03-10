"""
pages/notes.py — Notes & Évaluations
- Coef auto rempli selon le cours sélectionné
- Cours filtrés par niveau de l'étudiant (ex: ECO405 → 4e année)
- Bulletin PDF individuel beau (par étudiant, pas global)
- Import Excel robuste (accepte les colonnes du template)
- Export CSV / Excel / Stata / R
- Suppression avec confirmation
"""
import base64, io
from dash import html, dcc, callback, Input, Output, State, ctx, no_update, ALL
import pandas as pd
from models.database import get_db, Student, Course, Grade, Session
from utils.layout import page_header, C


def _get_base_data():
    db = get_db()
    courses  = db.query(Course).order_by(Course.code).all()
    students = db.query(Student).order_by(Student.nom).all()
    c_data   = [(c.code, c.libelle) for c in courses]
    s_data   = [(s.id, s.prenom, s.nom, s.annee or 3) for s in students]
    db.close()
    c_opts = [{"label":f"{code} — {lib}","value":code} for code,lib in c_data]
    s_opts = [{"label":f"{p} {n} ({a}e an.)","value":sid} for sid,p,n,a in s_data]
    return c_opts, s_opts, s_data


def _build_table(filt="all"):
    db = get_db()
    q  = db.query(Grade)
    if filt and filt != "all":
        q = q.filter(Grade.course_code==filt)
    grades = q.all()
    data = []
    for g in grades:
        data.append({"id":g.id,"etud":f"{g.student.prenom} {g.student.nom}" if g.student else str(g.id_student),
                     "code":g.course_code,"note":g.note,"coef":g.coefficient,"type":g.type_eval or "Examen"})
    db.close()
    if not data:
        return html.Div("Aucune note.", style={"padding":"24px","textAlign":"center",
                                               "color":C["muted"],"fontStyle":"italic"})
    rows = []
    for d in data:
        nc = C["success"] if d["note"]>=12 else C["warning"] if d["note"]>=8 else C["danger"]
        rows.append(html.Tr([
            html.Td(d["etud"],  style={"fontWeight":"700"}),
            html.Td(d["code"],  style={"color":C["bleu"],"fontWeight":"800"}),
            html.Td(html.Span(f"{d['note']}/20",
                              style={"background":f"{nc}18","color":nc,"padding":"3px 12px",
                                     "borderRadius":"20px","fontSize":"13px","fontWeight":"700"})),
            html.Td(f"x{d['coef']}"),
            html.Td(d["type"], style={"color":C["muted"],"fontSize":"13px"}),
            html.Td(html.Button("🗑", id={"type":"note-del-btn","index":d["id"]},
                                style={"background":"none","border":"none","cursor":"pointer",
                                       "color":"#C62828","fontSize":"15px"})),
        ]))
    return html.Table(className="sga-table", children=[
        html.Thead(html.Tr([html.Th("Étudiant"),html.Th("Cours"),html.Th("Note"),
                             html.Th("Coeff."),html.Th("Type"),html.Th("")])),
        html.Tbody(rows),
    ])


def _grades_to_df(course_code=None):
    db = get_db()
    q  = db.query(Grade)
    if course_code: q = q.filter(Grade.course_code==course_code)
    grades = q.all()
    rows = [{"id_student":g.id_student,"nom":g.student.nom if g.student else "",
             "prenom":g.student.prenom if g.student else "",
             "filiere":g.student.filiere if g.student else "",
             "cours":g.course_code,"libelle":g.course.libelle if g.course else "",
             "note":g.note,"coefficient":g.coefficient,"type_eval":g.type_eval} for g in grades]
    db.close()
    return pd.DataFrame(rows)


def layout():
    c_opts, s_opts, _ = _get_base_data()
    s_opts_bulletin   = [{"label":f"{p} {n}","value":sid} for sid,p,n,_ in _get_base_data()[2]]
    return html.Div([
        dcc.Store(id="note-del-store",     data=None),
        dcc.Store(id="note-pending-delete",data=None),
        page_header("Évaluations & Notes", "Saisie, imports, exports et bulletins individuels"),

        # ── Formulaire saisie ─────────────────────────────────────────────
        html.Div(style={"display":"flex","justifyContent":"center","marginBottom":"24px"}, children=[
            html.Div(style={"width":"100%","maxWidth":"700px"}, children=[
                html.Div(className="sga-card", children=[
                    html.Div("Saisir une note", className="sga-card-title"),
                    html.Div(style={"display":"grid","gridTemplateColumns":"1fr 1fr","gap":"16px"}, children=[
                        html.Div([html.Label("Étudiant *", className="form-label-sga"),
                                   dcc.Dropdown(id="n-etud", options=s_opts,
                                                placeholder="Sélectionner", style={"marginBottom":"0"})]),
                        html.Div([html.Label("Cours *", className="form-label-sga"),
                                   dcc.Dropdown(id="n-cours", options=c_opts,
                                                placeholder="Sélectionner", style={"marginBottom":"0"})]),
                    ]),
                    html.Div(style={"height":"14px"}),
                    html.Div(style={"display":"grid","gridTemplateColumns":"1fr 1fr 1fr","gap":"16px"}, children=[
                        html.Div([html.Label("Note (/20) *", className="form-label-sga"),
                                   dcc.Input(id="n-note", type="number", min=0, max=20, step=0.25,
                                             placeholder="14.50", className="sga-input",
                                             style={"marginBottom":"0"})]),
                        html.Div([html.Label("Coefficient (auto)", className="form-label-sga"),
                                   dcc.Input(id="n-coef", type="number", min=0.5, max=10, step=0.5,
                                             placeholder="1.0", className="sga-input",
                                             style={"marginBottom":"0","background":"#FFFDE7"})]),
                        html.Div([html.Label("Type", className="form-label-sga"),
                                   dcc.Dropdown(id="n-type",
                                                options=[{"label":t,"value":t} for t in
                                                         ["Examen","Partiel","Contrôle continu","Projet","Oral"]],
                                                value="Examen", style={"marginBottom":"0"})]),
                    ]),
                    html.Div(style={"height":"20px"}),
                    html.Button("Enregistrer la note", id="n-btn-save", className="btn-bleu",
                                style={"width":"100%","padding":"12px","fontSize":"15px"}),
                    html.Div(id="n-feedback", style={"marginTop":"10px"}),

                ]),
            ]),
        ]),

        html.Div(id="note-modal-overlay", style={"display":"none"}, children=[
            html.Div(className="modal-overlay", children=[
                html.Div(className="modal-box", children=[
                    html.Div("⚠", style={"fontSize":"36px","textAlign":"center","marginBottom":"10px","color":"#F9A825"}),
                    html.Div("Supprimer cette note ?", style={
                        "fontFamily":"'Merriweather',serif","fontSize":"19px","fontWeight":"900",
                        "color":"#0D47A1","textAlign":"center","marginBottom":"8px",
                    }),
                    html.Div(id="note-modal-label", style={
                        "textAlign":"center","fontWeight":"700","color":"#1565C0","fontSize":"14px","marginBottom":"14px",
                    }),
                    html.Div("Cette note sera définitivement supprimée du dossier de l'étudiant.", style={
                        "background":"#FFF8E1","border":"1px solid #FFE082","borderRadius":"8px",
                        "padding":"10px 14px","fontSize":"13px","color":"#5D4037","marginBottom":"22px",
                    }),
                    html.Div(style={"display":"flex","gap":"12px"}, children=[
                        html.Button("Annuler", id="note-confirm-cancel",
                                    style={"flex":"1","padding":"11px","background":"#EEF5FF",
                                           "color":"#0D47A1","border":"1.5px solid #BBDEFB",
                                           "borderRadius":"8px","cursor":"pointer","fontWeight":"700"}),
                        html.Button("Oui, supprimer", id="note-confirm-ok",
                                    style={"flex":"1","padding":"11px","background":"#C62828",
                                           "color":"#FFFFFF","border":"none","borderRadius":"8px",
                                           "cursor":"pointer","fontWeight":"700"}),
                    ]),
                ]),
            ]),
        ]),

        # ── Outils export / bulletin ──────────────────────────────────────
        html.Div(style={"display":"flex","justifyContent":"center","marginBottom":"24px"}, children=[
            html.Div(style={"width":"100%","maxWidth":"900px"}, children=[
                html.Div(className="sga-card", children=[
                    html.Div("Bulletin individuel & Exports", className="sga-card-title"),

                    html.Div(style={"display":"grid","gridTemplateColumns":"1fr 1fr","gap":"16px","marginBottom":"16px"}, children=[
                        html.Div([
                            html.Label("Étudiant (bulletin PDF)", className="form-label-sga"),
                            dcc.Dropdown(id="n-etud-bulletin", options=s_opts_bulletin,
                                         placeholder="Choisir un étudiant pour le bulletin",
                                         style={"marginBottom":"0"}),
                        ]),
                        html.Div([
                            html.Label("Cours (template / exports)", className="form-label-sga"),
                            dcc.Dropdown(id="n-cours-export", options=c_opts,
                                         placeholder="Optionnel — tous les cours si vide",
                                         style={"marginBottom":"0"}),
                        ]),
                    ]),

                    html.Div(style={"display":"grid","gridTemplateColumns":"1fr 1fr","gap":"12px","marginBottom":"12px"}, children=[
                        html.Button("📥 Template Excel (saisie notes)", id="n-btn-template",
                                    className="btn-bleu", style={"fontSize":"13px","padding":"11px"}),
                        html.Button("📄 Bulletin PDF individuel",       id="n-btn-pdf",
                                    className="btn-or",      style={"fontSize":"13px","padding":"11px"}),
                    ]),
                    html.Div(style={"display":"grid","gridTemplateColumns":"1fr 1fr 1fr 1fr","gap":"10px","marginBottom":"16px"}, children=[
                        html.Button("📊 CSV",   id="n-btn-csv",
                                    style={"background":"#F3F8FF","color":"#1565C0","border":"1px solid #90CAF9",
                                           "borderRadius":"8px","fontSize":"13px","padding":"10px","cursor":"pointer","fontWeight":"700"}),
                        html.Button("📗 Excel", id="n-btn-xlsx",
                                    style={"background":"#F1FBF4","color":"#2E7D32","border":"1px solid #A5D6A7",
                                           "borderRadius":"8px","fontSize":"13px","padding":"10px","cursor":"pointer","fontWeight":"700"}),
                        html.Button("📘 Stata", id="n-btn-stata",
                                    style={"background":"#EEF2FF","color":"#3730A3","border":"1px solid #A5B4FC",
                                           "borderRadius":"8px","fontSize":"13px","padding":"10px","cursor":"pointer","fontWeight":"700"}),
                        html.Button("📙 R/SAS", id="n-btn-r",
                                    style={"background":"#FFF7ED","color":"#C2410C","border":"1px solid #FCA974",
                                           "borderRadius":"8px","fontSize":"13px","padding":"10px","cursor":"pointer","fontWeight":"700"}),
                    ]),
                    html.Div(id="n-export-status"),

                    html.Div(style={"borderTop":f"1px solid {C['bordure']}","paddingTop":"16px"}, children=[
                        html.Div("Importer un template Excel rempli", style={
                            "fontWeight":"700","fontSize":"14px","fontFamily":"'EB Garamond',serif",
                            "color":C["bleu3"],"marginBottom":"10px",
                        }),
                        html.Div([
                            html.Div("📌 Colonnes attendues dans votre fichier : ", style={"fontSize":"12px","color":C["muted"],"display":"inline"}),
                            html.Code("ID, Nom, Prenom, Filiere, NOTE (/20), Coefficient, Type",
                                      style={"fontSize":"11px","background":"#F5E6C8","padding":"2px 8px","borderRadius":"4px"}),
                        ], style={"marginBottom":"10px"}),
                        dcc.Upload(id="n-upload",
                                   children=html.Div([
                                       html.Span("📂", style={"fontSize":"36px"}),
                                       html.Div("Glissez-déposez ou cliquez",
                                                style={"fontSize":"13px","color":C["muted"],"marginTop":"8px"}),
                                       html.Div(".xlsx uniquement",
                                                style={"fontSize":"11px","color":C["muted"]}),
                                   ]),
                                   className="upload-zone", accept=".xlsx"),
                        html.Div(id="n-upload-fb"),
                    ]),
                ]),
            ]),
        ]),

        # ── Tableau notes ─────────────────────────────────────────────────
        html.Div(className="sga-card", style={"padding":"0"}, children=[
            html.Div(style={"display":"flex","justifyContent":"space-between","alignItems":"center",
                            "padding":"20px 20px 14px"}, children=[
                html.Div("Tableau des notes", className="sga-card-title",
                         style={"marginBottom":"0","borderBottom":"none","paddingBottom":"0"}),
                dcc.Dropdown(id="n-filter",
                             options=[{"label":"Tous les cours","value":"all"}]+c_opts,
                             value="all", clearable=False, style={"width":"280px"}),
            ]),
            html.Div(id="n-table", children=_build_table("all")),
        ]),

        dcc.Download(id="n-dl-template"),
        dcc.Download(id="n-dl-pdf"),
        dcc.Download(id="n-dl-csv"),
        dcc.Download(id="n-dl-xlsx"),
        dcc.Download(id="n-dl-stata"),
        dcc.Download(id="n-dl-r"),
    ])


# ── Auto-coef quand cours sélectionné ────────────────────────────────────────
@callback(
    Output("n-coef","value"),
    Input("n-cours","value"),
    prevent_initial_call=True,
)
def auto_coef(course_code):
    """Remplit le coef avec celui enregistré pour ce cours."""
    if not course_code: return no_update
    db = get_db()
    c  = db.query(Course).filter(Course.code==course_code).first()
    db.close()
    # On utilise le champ description pour stocker le coef — sinon on déduit du code
    # Convention : si le code contient un chiffre à la fin ex STAT301 → année 3 → coef 1.5
    if not c: return no_update
    # Déduire le coef selon le niveau dans le code (ex: 301→3e an, 201→2e an, 401→4e an)
    import re
    m = re.search(r'(\d)(\d{2})$', c.code)
    if m:
        niveau = int(m.group(1))
        coef_map = {1: 1.0, 2: 1.5, 3: 2.0, 4: 2.5}
        return coef_map.get(niveau, 1.0)
    return 1.0


# ── Enregistrer note + confirmation suppression ───────────────────────────────
@callback(
    Output("n-feedback",     "children"),
    Output("n-table",        "children"),
    Output("note-modal-overlay","style"),
    Output("note-modal-label",  "children"),
    Output("note-pending-delete","data"),
    Input("n-btn-save",          "n_clicks"),
    Input("n-filter",            "value"),
    Input("note-del-store",      "data"),
    Input("note-confirm-ok",     "n_clicks"),
    Input("note-confirm-cancel", "n_clicks"),
    State("n-cours","value"), State("n-etud","value"),
    State("n-note","value"),  State("n-coef","value"),
    State("n-type","value"),  State("note-pending-delete","data"),
    prevent_initial_call=True,
)
def handle_note(n_save, filt, del_id, n_ok, n_cancel,
                code, stud_id, note_val, coef, type_eval, pending):
    t = ctx.triggered_id

    if t == "note-del-store" and del_id:
        db = get_db()
        g  = db.query(Grade).filter(Grade.id==del_id).first()
        label = f"{g.student.prenom} {g.student.nom} — {g.course_code} — {g.note}/20" if g else str(del_id)
        db.close()
        return (no_update, no_update, {"display":"block"}, label, del_id)

    if t == "note-confirm-cancel":
        return (html.Div("Suppression annulée.", className="alert-sga",
                         style={"background":"#E3F2FD","color":"#1565C0","borderLeftColor":"#1565C0"}),
                no_update, {"display":"none"}, "", None)

    if t == "note-confirm-ok" and pending:
        db = get_db()
        g  = db.query(Grade).filter(Grade.id==pending).first()
        if g: db.delete(g); db.commit()
        db.close()
        return (html.Div("Note supprimée.", className="alert-sga",
                         style={"background":"#FFF3E0","color":C["warning"],"borderLeftColor":C["warning"]}),
                _build_table(filt or "all"), {"display":"none"}, "", None)

    if t == "n-filter":
        return no_update, _build_table(filt or "all"), no_update, no_update, no_update

    if t == "n-btn-save" and n_save:
        if not code or stud_id is None or note_val is None:
            return (html.Div("Étudiant, cours et note sont requis.", className="alert-sga",
                             style={"background":"#FFEBEE","color":C["danger"],"borderLeftColor":C["danger"]}),
                    no_update, no_update, no_update)
        db = get_db()
        ex = db.query(Grade).filter(Grade.id_student==stud_id, Grade.course_code==code).first()
        if ex:
            ex.note=float(note_val); ex.coefficient=float(coef or 1); ex.type_eval=type_eval or "Examen"
        else:
            db.add(Grade(id_student=int(stud_id), course_code=code,
                         note=float(note_val), coefficient=float(coef or 1),
                         type_eval=type_eval or "Examen"))
        db.commit(); db.close()
        return (html.Div("Note enregistrée.", className="alert-sga",
                         style={"background":"#E8F5E9","color":C["success"],"borderLeftColor":C["success"]}),
                _build_table(filt or "all"), {"display":"none"}, "", None)

    return no_update, _build_table(filt or "all"), no_update, no_update, no_update


@callback(Output("note-del-store","data"),
          Input({"type":"note-del-btn","index":ALL},"n_clicks"), prevent_initial_call=True)
def cap_note_del(clicks):
    if not any(n for n in clicks if n): return no_update
    t = ctx.triggered_id
    return t["index"] if isinstance(t, dict) else no_update


# ── Template Excel ────────────────────────────────────────────────────────────
@callback(
    Output("n-dl-template",   "data"),
    Output("n-export-status", "children", allow_duplicate=True),
    Input("n-btn-template",   "n_clicks"),
    State("n-cours-export",   "value"),
    prevent_initial_call=True,
)
def dl_template(n, course_code):
    if not n: return no_update, no_update
    if not course_code:
        return no_update, html.Div("Sélectionnez un cours pour le template.", className="alert-sga",
                                   style={"background":"#FFF3E0","color":C["warning"],"borderLeftColor":C["warning"]})
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        db       = get_db()
        course   = db.query(Course).filter(Course.code==course_code).first()
        students = db.query(Student).order_by(Student.nom).all()
        s_rows   = [(s.id, s.nom, s.prenom, s.filiere or "") for s in students]
        db.close()

        wb = Workbook(); ws = wb.active; ws.title = course_code[:31]
        BDX = PatternFill("solid",fgColor="0D47A1")
        OR  = PatternFill("solid",fgColor="C8973A")
        CRM = PatternFill("solid",fgColor="FDF8F0")
        ALT = PatternFill("solid",fgColor="FFF8EC")
        WHT = PatternFill("solid",fgColor="FFFFFF")
        thin= Side(style="thin",color="E8D5B7")
        BDR = Border(left=thin,right=thin,top=thin,bottom=thin)
        CTR = Alignment(horizontal="center",vertical="center")
        LFT = Alignment(horizontal="left",  vertical="center")

        ws.merge_cells("A1:G1")
        ws["A1"] = f"ENSAE Dakar — Notes — {course_code}" + (f" : {course.libelle}" if course else "")
        ws["A1"].font=Font(bold=True,color="FFFFFF",name="Calibri",size=14)
        ws["A1"].fill=BDX; ws["A1"].alignment=CTR; ws.row_dimensions[1].height=32

        ws.merge_cells("A2:G2")
        ws["A2"] = "Remplissez la colonne NOTE (/20). Ne modifiez pas ID, Nom, Prenom, Filiere."
        ws["A2"].font=Font(italic=True,color="9A7B6A",name="Calibri",size=9)
        ws["A2"].fill=CRM; ws["A2"].alignment=CTR; ws.row_dimensions[2].height=16

        # En-têtes — EXACTEMENT les noms attendus par l'import
        headers=["ID","Nom","Prenom","Filiere","NOTE (/20)","Coefficient","Type"]
        widths =[8,20,18,16,14,13,18]
        for col,(h,w) in enumerate(zip(headers,widths),1):
            c=ws.cell(row=3,column=col,value=h)
            c.font=Font(bold=True,color="FFFFFF",name="Calibri",size=10)
            c.fill=OR if col==5 else BDX
            c.alignment=CTR; c.border=BDR
            ws.column_dimensions[get_column_letter(col)].width=w
        ws.row_dimensions[3].height=22

        for i,(sid,nom,prenom,filiere) in enumerate(s_rows):
            r    = i+4
            fill = ALT if i%2==0 else WHT
            for col,val in enumerate([sid,nom,prenom,filiere,"",1.0,"Examen"],1):
                c=ws.cell(row=r,column=col,value=val)
                c.fill=fill; c.border=BDR
                c.font=Font(bold=True,color="0D47A1",name="Calibri",size=10) if col==5 else Font(name="Calibri",size=10)
                c.alignment=CTR if col in (1,5,6) else LFT
            ws.row_dimensions[r].height=18

        ws.freeze_panes="A4"
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return (dcc.send_bytes(buf.read(),filename=f"template_notes_{course_code}.xlsx"),
                html.Div("Template généré avec succès.", className="alert-sga",
                         style={"background":"#E8F5E9","color":C["success"],"borderLeftColor":C["success"]}))
    except Exception as e:
        return no_update, html.Div(f"Erreur: {e}", className="alert-sga",
                                   style={"background":"#FFEBEE","color":C["danger"],"borderLeftColor":C["danger"]})


# ── Import Excel robuste ──────────────────────────────────────────────────────
@callback(
    Output("n-upload-fb","children"),
    Input("n-upload","contents"),
    State("n-upload","filename"),
    State("n-cours-export","value"),
    prevent_initial_call=True,
)
def import_excel(contents, filename, course_code):
    if not contents: return no_update
    if not course_code:
        return html.Div("Sélectionnez un cours avant d'importer.", className="alert-sga",
                        style={"background":"#FFF3E0","color":C["warning"],"borderLeftColor":C["warning"]})
    try:
        _, b64 = contents.split(",")
        raw_bytes = io.BytesIO(base64.b64decode(b64))

        # ── Détection automatique de la ligne de header ──────────────────
        # Cherche la 1ère ligne qui contient "ID" dans une cellule
        import unicodedata, openpyxl
        def norm(s):
            s = str(s).lower().strip()
            return "".join(ch for ch in unicodedata.normalize("NFD", s)
                           if unicodedata.category(ch) != "Mn")

        wb_peek = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)), data_only=True, read_only=True)
        ws_peek = wb_peek.active
        header_row = 0  # 0-indexed pour pandas header=
        for i, row in enumerate(ws_peek.iter_rows(max_row=10, values_only=True)):
            row_norms = [norm(str(v)) for v in row if v is not None]
            if any(v in ("id", "id_student", "num", "numero") for v in row_norms):
                header_row = i
                break
        wb_peek.close()

        raw_bytes.seek(0)
        df = pd.read_excel(raw_bytes, header=header_row)

        col_map = {norm(str(c)): c for c in df.columns}

        # Trouver colonne ID
        id_col = None
        for cand in ("id","id_student","student_id","etudiant_id","num","numero"):
            if cand in col_map: id_col = col_map[cand]; break

        # Trouver colonne NOTE
        note_col = None
        for cand in ("note (/20)","note(/20)","note","grade","score","note_finale","resultat"):
            if cand in col_map: note_col = col_map[cand]; break

        # Trouver colonne COEF
        coef_col = None
        for cand in ("coefficient","coef","coeff","poids"):
            if cand in col_map: coef_col = col_map[cand]; break

        if id_col is None:
            return html.Div(
                f"Colonne ID introuvable (ligne détectée: {header_row+1}). "
                f"Colonnes lues: {', '.join(str(c) for c in df.columns[:8])}",
                className="alert-sga",
                style={"background":"#FFEBEE","color":C["danger"],"borderLeftColor":C["danger"]})

        if note_col is None:
            return html.Div(
                f"Colonne NOTE introuvable. Colonnes: {', '.join(str(c) for c in df.columns[:8])}",
                className="alert-sga",
                style={"background":"#FFEBEE","color":C["danger"],"borderLeftColor":C["danger"]})

        db = get_db(); cnt = 0; skipped = 0
        for _, row in df.iterrows():
            note_raw = row.get(note_col)
            if pd.isna(note_raw): skipped += 1; continue
            try:
                sid  = int(row[id_col])
                coef = float(row[coef_col]) if (coef_col and pd.notna(row.get(coef_col))) else 1.0
                ex   = db.query(Grade).filter(Grade.id_student==sid, Grade.course_code==course_code).first()
                if ex: ex.note=float(note_raw); ex.coefficient=coef
                else:  db.add(Grade(id_student=sid,course_code=course_code,
                                    note=float(note_raw),coefficient=coef))
                cnt += 1
            except Exception as e: skipped += 1; continue
        db.commit(); db.close()

        msg = f"{cnt} note(s) importée(s)"
        if skipped: msg += f" ({skipped} ligne(s) ignorée(s))"
        return html.Div(msg + ".", className="alert-sga",
                        style={"background":"#E8F5E9","color":C["success"],"borderLeftColor":C["success"]})
    except Exception as e:
        return html.Div(f"Erreur import : {e}", className="alert-sga",
                        style={"background":"#FFEBEE","color":C["danger"],"borderLeftColor":C["danger"]})


# ── Bulletin PDF INDIVIDUEL ───────────────────────────────────────────────────
@callback(
    Output("n-dl-pdf",        "data"),
    Output("n-export-status", "children", allow_duplicate=True),
    Input("n-btn-pdf",        "n_clicks"),
    State("n-etud-bulletin",  "value"),
    prevent_initial_call=True,
)
def gen_bulletin(n, stud_id):
    if not n: return no_update, no_update
    if not stud_id:
        return no_update, html.Div("Sélectionnez un étudiant pour le bulletin.", className="alert-sga",
                                   style={"background":"#FFF3E0","color":C["warning"],"borderLeftColor":C["warning"]})
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                         Paragraph, Spacer, HRFlowable)
        from reportlab.lib.units import cm
        from datetime import datetime

        db     = get_db()
        s      = db.query(Student).filter(Student.id==stud_id).first()
        if not s: db.close(); return no_update, no_update

        grades     = s.grades
        nb_abs     = len(s.attendances)
        total_sess = db.query(Session).count()

        # Heures d'absence par matière
        abs_par_cours = {}
        for att in s.attendances:
            code = att.session.course_code if att.session else None
            if code:
                abs_par_cours[code] = abs_par_cours.get(code, 0) + (att.session.duree or 1.5)

        moy = (sum(g.note*g.coefficient for g in grades)/sum(g.coefficient for g in grades)) if grades else None
        taux = round(nb_abs/total_sess*100) if total_sess else 0
        mention = ("Très Bien" if moy and moy>=16 else "Bien" if moy and moy>=14
                   else "Assez Bien" if moy and moy>=12 else "Passable" if moy and moy>=10
                   else "Insuffisant") if moy else "—"
        dob_str      = s.date_naissance.strftime("%d/%m/%Y") if s.date_naissance else "—"
        niveau_label = {1:"1ère année",2:"2e année",3:"3e année",4:"4e année"}.get(s.annee,f"{s.annee}e an.")
        g_data = [(g.course_code, g.course.libelle if g.course else "",
                   g.note, g.coefficient, g.type_eval,
                   abs_par_cours.get(g.course_code, 0.0)) for g in grades]
        db.close()

        BDX = colors.HexColor("#0D47A1"); OR  = colors.HexColor("#F9A825")
        CRM = colors.HexColor("#E3F2FD"); MUT = colors.HexColor("#607D8B")
        GRN = colors.HexColor("#2E7D32"); RED = colors.HexColor("#C62828")
        ORG = colors.HexColor("#E65100")

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf,pagesize=A4,leftMargin=2*cm,rightMargin=2*cm,
                                  topMargin=2*cm,bottomMargin=2*cm)
        def ps(name,**kw): return ParagraphStyle(name,**kw)

        story = [
            # En-tête
            Paragraph("ENSAE Dakar",   ps("h1",fontName="Helvetica-Bold",fontSize=24,textColor=BDX,spaceAfter=2)),
            Paragraph("Bulletin de Notes Individuel",
                       ps("h2",fontName="Helvetica",fontSize=14,textColor=MUT,spaceAfter=2)),
            Paragraph(f"Émis le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
                       ps("d",fontName="Helvetica",fontSize=9,textColor=MUT,spaceAfter=6)),
            HRFlowable(width="100%",thickness=3,color=OR),
            Spacer(1,0.4*cm),

            # Identité
            Paragraph("Identité de l'Étudiant",
                       ps("sec",fontName="Helvetica-Bold",fontSize=11,textColor=BDX,spaceAfter=5)),
        ]

        id_data = [
            ["Nom & Prénom", f"{s.prenom} {s.nom}", "N° Étudiant", str(s.id)],
            ["Email",        s.email,                "Filière",     s.filiere or "—"],
            ["Date de naissance", dob_str,           "Niveau",      niveau_label],
        ]
        it = Table(id_data,colWidths=[3.5*cm,6*cm,3*cm,4*cm])
        it.setStyle(TableStyle([
            ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),
            ("TEXTCOLOR",(0,0),(0,-1),BDX),("TEXTCOLOR",(2,0),(2,-1),BDX),
            ("FONTSIZE",(0,0),(-1,-1),9),("BACKGROUND",(0,0),(-1,-1),CRM),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#CFE2F3")),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
        ]))
        story += [it, Spacer(1,0.35*cm)]

        # Synthèse
        moy_c = GRN if (moy and moy>=12) else ORG if (moy and moy>=8) else RED
        abs_c = RED if taux>20 else ORG if taux>10 else GRN
        syn = [["Moyenne Générale","Absentéisme","Mention","Matières évaluées"],
               [f"{moy:.2f}/20" if moy else "—",f"{taux}% ({nb_abs} séances)",mention,str(len(g_data))]]
        st = Table(syn,colWidths=[4*cm,4*cm,4.5*cm,4*cm])
        st.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),BDX),("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),9),
            ("BACKGROUND",(0,1),(-1,-1),CRM),
            ("TEXTCOLOR",(0,1),(0,1),moy_c),("TEXTCOLOR",(1,1),(1,1),abs_c),
            ("FONTNAME",(0,1),(-1,-1),"Helvetica-Bold"),("FONTSIZE",(0,1),(-1,-1),13),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8),
            ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#CFE2F3")),
        ]))
        story += [st, Spacer(1,0.4*cm),
                  Paragraph("Détail par Matière",
                             ps("sec2",fontName="Helvetica-Bold",fontSize=11,textColor=BDX,spaceAfter=5))]

        if g_data:
            td = [["Code","Matière","Note /20","Coeff.","Type","H. Abs.","Mention"]]
            for code,lib,note,coef,typ,h_abs in g_data:
                m2 = "TB" if note>=16 else "B" if note>=14 else "AB" if note>=12 else "P" if note>=10 else "I"
                nc = GRN if note>=12 else ORG if note>=8 else RED
                td.append([code,lib[:30],f"{note:.2f}",str(coef),typ or "Examen",f"{h_abs:.1f}h",m2])
            t2 = Table(td,colWidths=[2*cm,5.5*cm,2.2*cm,1.8*cm,2.8*cm,1.8*cm,1.9*cm])
            t2.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),BDX),("TEXTCOLOR",(0,0),(-1,0),colors.white),
                ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),9),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[CRM,colors.white]),
                ("FONTSIZE",(0,1),(-1,-1),9),
                ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#CFE2F3")),
                ("ALIGN",(2,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
            ]))
            story.append(t2)
        else:
            story.append(Paragraph("Aucune note enregistrée.",
                                    ps("e",fontName="Helvetica",fontSize=10,textColor=MUT)))

        story += [
            Spacer(1,0.5*cm),
            HRFlowable(width="100%",thickness=1,color=colors.HexColor("#CFE2F3")),
            Paragraph("Gilbert OUMSAORE & Josée JEAZE — ENSAE Dakar — Data Visualisation 2025-2026",
                       ps("f",fontName="Helvetica",fontSize=8,textColor=MUT,spaceBefore=4)),
        ]
        doc.build(story); buf.seek(0)
        fname = f"bulletin_{s.nom}_{s.prenom}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return (dcc.send_bytes(buf.read(),filename=fname),
                html.Div(f"Bulletin de {s.prenom} {s.nom} généré.", className="alert-sga",
                         style={"background":"#E8F5E9","color":C["success"],"borderLeftColor":C["success"]}))
    except ImportError:
        return no_update, html.Div("Installez reportlab : pip install reportlab", className="alert-sga",
                                   style={"background":"#FFF3E0","color":C["warning"],"borderLeftColor":C["warning"]})
    except Exception as e:
        return no_update, html.Div(f"Erreur PDF : {e}", className="alert-sga",
                                   style={"background":"#FFEBEE","color":C["danger"],"borderLeftColor":C["danger"]})


# ── Exports CSV / Excel / Stata / R ──────────────────────────────────────────
@callback(Output("n-dl-csv","data"), Output("n-export-status","children",allow_duplicate=True),
          Input("n-btn-csv","n_clicks"), State("n-cours-export","value"), prevent_initial_call=True)
def exp_csv(n,code):
    if not n: return no_update,no_update
    df=_grades_to_df(code); buf=io.StringIO(); df.to_csv(buf,index=False,encoding="utf-8-sig"); buf.seek(0)
    return (dcc.send_string(buf.read(),filename=f"notes_{code or 'all'}.csv"),
            html.Div("Export CSV généré.",className="alert-sga",style={"background":"#E8F5E9","color":C["success"],"borderLeftColor":C["success"]}))


@callback(Output("n-dl-xlsx","data"), Output("n-export-status","children",allow_duplicate=True),
          Input("n-btn-xlsx","n_clicks"), State("n-cours-export","value"), prevent_initial_call=True)
def exp_xlsx(n,code):
    if not n: return no_update,no_update
    df=_grades_to_df(code); buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine="openpyxl") as w: df.to_excel(w,index=False,sheet_name="Notes")
    buf.seek(0)
    return (dcc.send_bytes(buf.read(),filename=f"notes_{code or 'all'}.xlsx"),
            html.Div("Export Excel généré.",className="alert-sga",style={"background":"#E8F5E9","color":C["success"],"borderLeftColor":C["success"]}))


@callback(Output("n-dl-stata","data"), Output("n-export-status","children",allow_duplicate=True),
          Input("n-btn-stata","n_clicks"), State("n-cours-export","value"), prevent_initial_call=True)
def exp_stata(n,code):
    if not n: return no_update,no_update
    try:
        df=_grades_to_df(code); buf=io.BytesIO(); df.to_stata(buf,write_index=False); buf.seek(0)
        return (dcc.send_bytes(buf.read(),filename=f"notes_{code or 'all'}.dta"),
                html.Div("Export Stata (.dta) généré.",className="alert-sga",style={"background":"#E8F5E9","color":C["success"],"borderLeftColor":C["success"]}))
    except Exception as e:
        return no_update,html.Div(f"Erreur Stata: {e}",className="alert-sga",style={"background":"#FFEBEE","color":C["danger"],"borderLeftColor":C["danger"]})


@callback(Output("n-dl-r","data"), Output("n-export-status","children",allow_duplicate=True),
          Input("n-btn-r","n_clicks"), State("n-cours-export","value"), prevent_initial_call=True)
def exp_r(n,code):
    if not n: return no_update,no_update
    df=_grades_to_df(code); buf=io.StringIO()
    buf.write("# Compatible R/SAS — charger avec: df <- read.csv2('fichier.csv')\n")
    df.to_csv(buf,index=False,sep=";"); buf.seek(0)
    return (dcc.send_string(buf.read(),filename=f"notes_{code or 'all'}_R.csv"),
            html.Div("Export R/SAS généré.",className="alert-sga",style={"background":"#E8F5E9","color":C["success"],"borderLeftColor":C["success"]}))
